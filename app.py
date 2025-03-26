# -*- coding: utf-8 -*-
from flask import Flask, request, jsonify
import pandas as pd
from sklearn.preprocessing import LabelEncoder
from geopy.geocoders import Nominatim
import os
import json
import re
import pandas as pd
from sklearn.preprocessing import LabelEncoder
import matplotlib.pyplot as plt
import seaborn as sns


app = Flask(__name__)

@app.route("/")
def home():
    return "Welcome to the Risk Rate API!"

def normalize(text):
    return str(text).lower().strip().replace("-", "").replace(" ", "")





file_url = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRnH-_FyYUJ5NCF_HHQjT1JhCGl7MsMxRlsRWVib3wi7P78LHuDgkLk2RwjlcuXNQ/pub?output=csv"

try:
    # Attempt to read the file
    df = pd.read_csv(file_url)
    print("CSV file loaded successfully.")
except Exception as e:
    print(f"Error loading CSV: {e}")
    exit()

# Select only the required columns
columns_to_extract = ['LicenseStatus', 'ActivityMainGroup', 'MainActivity', 'Street', 'Region', 'LastApplicationNo', 'Numberofrenewals']
df = df[columns_to_extract]

# Define the labeling logic
def label_row(row):
    if row['Numberofrenewals'] >= 3 and row['LicenseStatus'] in ['Active', 'Expired']:
        return 'high success'
    elif row['Numberofrenewals'] < 3 and row['LicenseStatus'] in ['Active', 'Expired']:
        return 'low success'
    elif row['Numberofrenewals'] >= 3 and row['LicenseStatus'] in ['Canceled', 'Totally Blocked', 'Managerially Canceled', 'Mortgaged', 'Partially Blocked', 'Suspended']:
        return 'good cancel'
    elif row['Numberofrenewals'] < 3 and row['LicenseStatus'] in ['Canceled', 'Totally Blocked', 'Managerially Canceled', 'Mortgaged', 'Partially Blocked', 'Suspended']:
        return 'bad cancel'
    elif row['LicenseStatus'] == 'Inactive':
        return 'new'
    else:
        return 'other'  # Optional: To handle cases that don't match any condition

# Apply the labeling function
df['label'] = df.apply(label_row, axis=1)

# Display the updated DataFrame
print(df[['LicenseStatus', 'Numberofrenewals', 'label']])  # Showing only relevant columns for clarity







clustering_columns = ['MainActivity', 'Region']
df_cluster = df[clustering_columns].copy()

# Step 2: Convert categorical variables to numerical using Label Encoding
label_encoders = {}
for col in clustering_columns:
    le = LabelEncoder()
    df_cluster[col] = le.fit_transform(df_cluster[col])  # Convert text to numbers
    label_encoders[col] = le  # Store encoders for later decoding if needed


# Step 3: Assign a unique cluster number for each unique MainActivity
df_cluster['Cluster'] = df_cluster.groupby('MainActivity').ngroup()

# Step 4: Merge clustering results back to the original dataset
df['Cluster'] = df_cluster['Cluster']

# Step 5: Visualize the Clusters
plt.figure(figsize=(8, 6))
sns.scatterplot(x=df_cluster['MainActivity'], y=df_cluster['Region'], hue=df_cluster['Cluster'], palette='viridis', s=100)
plt.title('Unique Clustering of MainActivity')
plt.xlabel('MainActivity (Encoded)')
plt.ylabel('Region (Encoded)')
plt.show()

# Step 6: Display results
print(df[['MainActivity', 'Region', 'Cluster']])






import pandas as pd

# Define label values
label_values = {
    'high success': 4,
    'low success': 3,
    'good cancel': 2,
    'bad cancel': 1,
    'new': 0
}

# Group by MainActivity and Region, count the occurrences of each label
risk_df = df.groupby(['MainActivity', 'Region', 'label']).size().unstack(fill_value=0)


# Ensure all label categories exist
for label in label_values.keys():
    if label not in risk_df.columns:
        risk_df[label] = 0

# Calculate risk score using the given formula
risk_df['RiskRate'] =   ( 1 - ((
    risk_df['high success'] * 4 +
    risk_df['low success'] * 3 +
    risk_df['good cancel'] * 2 +
    risk_df['bad cancel'] * 1 +
    risk_df['new'] * 0
) / (100 * 4)) ) * 100


# Merge back into original DataFrame
df = df.merge(risk_df[['RiskRate']], on=['MainActivity', 'Region'], how='left')

# Display results
print(df[['MainActivity', 'Region', 'RiskRate']])


df_clean = df.dropna()
print(df_clean[['MainActivity', 'Region', 'RiskRate']])



# Convert to DataFrame
df = pd.DataFrame(df_clean)
# Save to Excel file
df.to_excel("output2.xlsx", index=False, engine="openpyxl")


print("Excel file saved successfully!")



from openpyxl import load_workbook

# Load the workbook and select the sheet
file_path = "output2.xlsx"
wb = load_workbook(file_path)
ws = wb.active  # Select the first sheet

# Iterate over cells and replace negative values
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, (int, float)) and cell.value < 0:
            cell.value = 0


@app.route("/get_risk_rate", methods=["GET"])
def get_risk_rate():
    latitude = request.args.get("latitude", type=float)
    longitude = request.args.get("longitude", type=float)

    try:
        print(f"Request: lat={latitude}, lng={longitude}")
        region = get_region_from_latlng(latitude, longitude)
        print(f"Resolved region: {region}")

        risk_data = process_data()

        for entry in risk_data:
            if normalize(entry['Region']) == normalize(region):
                print(f"Matched region: {entry['Region']} => RiskRate: {entry['RiskRate']}")
                return jsonify({
                    "latitude": latitude,
                    "longitude": longitude,
                    "risk_rate": entry['RiskRate']
                })

        print("No match found, returning risk_rate: 0")
        return jsonify({
            "latitude": latitude,
            "longitude": longitude,
            "risk_rate": 0
        })

    except Exception as e:
        print(f"🔥 Error in /get_risk_rate: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/process", methods=['GET'])
def get_data():
    result = process_data()
    return jsonify(result)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
